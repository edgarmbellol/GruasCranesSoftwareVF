from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, Equipo, TipoEquipo, Marca, EstadoEquipo, Cargo, Cliente, RegistroHoras
from forms import UserForm, UserEditForm, UserSearchForm, EquipoForm, EquipoSearchForm, TipoEquipoForm, MarcaForm, EstadoEquipoForm, CargoForm, ClienteForm, ClienteSearchForm, RegistroHorasForm, CambiarContrasenaForm
import os
import shutil
import hashlib
from datetime import timedelta, datetime, date
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from functools import wraps
import qrcode
from PIL import Image
from dateutil.relativedelta import relativedelta
import calendar
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

def get_colombia_datetime():
    """Obtiene la fecha y hora actual de Colombia"""
    colombia_tz = pytz.timezone('America/Bogota')
    return datetime.now(colombia_tz)

def get_file_version(file_path):
    """Genera una versión única basada en el hash del archivo para evitar problemas de caché"""
    try:
        if os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                file_hash = hashlib.md5(f.read()).hexdigest()
                return file_hash[:8]  # Solo los primeros 8 caracteres
    except:
        pass
    return str(int(datetime.now().timestamp()))  # Fallback a timestamp

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'tu-clave-secreta-muy-segura-aqui')
app.permanent_session_lifetime = timedelta(hours=24)

# Hacer las funciones disponibles en todos los templates
@app.context_processor
def inject_functions():
    def format_colombia_datetime(dt):
        """Convertir datetime a zona horaria de Colombia"""
        if dt is None:
            return None
        if dt.tzinfo is None:
            # Si no tiene zona horaria, asumir que es UTC
            colombia_tz = pytz.timezone('America/Bogota')
            dt = pytz.utc.localize(dt)
        return dt.astimezone(pytz.timezone('America/Bogota'))
    
    return {
        'get_file_version': get_file_version,
        'get_colombia_datetime': get_colombia_datetime,
        'format_colombia_datetime': format_colombia_datetime
    }

# Configuración de base de datos
# Configuración PostgreSQL
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_PORT = os.environ.get('DB_PORT', '5432')
DB_NAME = os.environ.get('DB_NAME', 'gruas_db')
DB_USER = os.environ.get('DB_USER', 'gruas_user')
DB_PASSWORD = os.environ.get('DB_PASSWORD', 'gruas_password')

# Construir URL de conexión
from urllib.parse import quote_plus
app.config['SQLALCHEMY_DATABASE_URI'] = f"postgresql://{DB_USER}:{quote_plus(DB_PASSWORD)}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

# Configuraciones de rendimiento
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 20,           # Número de conexiones en el pool
    'max_overflow': 30,        # Conexiones adicionales permitidas
    'pool_timeout': 30,        # Tiempo de espera para conexión
    'pool_recycle': 3600,      # Reciclar conexiones cada hora
    'pool_pre_ping': True,     # Verificar conexiones antes de usar
    'echo': False              # Cambiar a True para debug de SQL
}
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Inicializar base de datos
db.init_app(app)

# Función para generar códigos QR
def generar_qr_equipo(equipo_id, placa):
    """Genera un código QR para un equipo específico"""
    try:
        # Crear directorio para QR si no existe
        qr_dir = os.path.join(app.static_folder, 'qr_codes')
        if not os.path.exists(qr_dir):
            os.makedirs(qr_dir)
        
        # URL del panel del equipo (nueva funcionalidad)
        # Usar BASE_URL del entorno si está disponible, sino usar URL fija
        base_url = os.environ.get('BASE_URL', 'https://gestor.gruascranes.com')
        url_panel = f"{base_url}/equipo/{equipo_id}"
        
        # DEBUG: Agregar prints para verificar la URL
        print(f"🔍 DEBUG QR: equipo_id={equipo_id}, base_url={base_url}")
        print(f"🔍 DEBUG QR: url_panel={url_panel}")
        
        # Crear código QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(url_panel)
        qr.make(fit=True)
        
        # Crear imagen del QR
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Guardar imagen
        filename = f"qr_equipo_{equipo_id}_{placa}.png"
        filepath = os.path.join(qr_dir, filename)
        img.save(filepath)
        
        print(f"✅ DEBUG QR: QR guardado en {filepath}")
        
        # Sincronizar con directorio de producción si existe
        sync_to_production(filepath, filename)
        
        return f"qr_codes/{filename}"
    except Exception as e:
        print(f"Error generando QR para equipo {equipo_id}: {str(e)}")
        return None

def sync_to_production(filepath, filename):
    """Registra archivo QR para sincronización con producción"""
    try:
        # Construir la ruta completa del archivo
        full_filepath = os.path.join(app.static_folder, filepath)
        
        if os.path.exists(full_filepath):
            # Solo registrar que se necesita sincronización
            # La sincronización real se hará por un proceso externo
            print(f"📝 QR generado localmente: {filename} - Pendiente de sincronización")
            
            # Crear un archivo de marcador para indicar que hay QR pendientes de sincronizar
            sync_marker = os.path.join(app.static_folder, 'qr_codes', '.sync_pending')
            with open(sync_marker, 'a') as f:
                f.write(f"{filename}\n")
                
        else:
            print(f"⚠️  Archivo QR no encontrado: {full_filepath}")
    except Exception as e:
        print(f"⚠️  Error registrando QR para sincronización: {e}")

def regenerar_todos_qr():
    """Regenera todos los códigos QR para equipos existentes"""
    equipos = Equipo.query.filter_by(Estado='activo').all()
    qr_generados = 0
    errores = 0
    
    for equipo in equipos:
        qr_path = generar_qr_equipo(equipo.IdEquipo, equipo.Placa)
        if qr_path:
            qr_generados += 1
        else:
            errores += 1
    
    # Ejecutar sincronización completa con el script Python
    try:
        import subprocess
        result = subprocess.run(['python3', '/home/mauricio/apps/flask_app/sync_qr_auto.py'], 
                              capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            print("✅ Sincronización completa ejecutada exitosamente")
        else:
            print(f"⚠️  Error en sincronización completa: {result.stderr}")
    except subprocess.TimeoutExpired:
        print("⏰ Timeout en sincronización completa")
    except Exception as e:
        print(f"⚠️  Error ejecutando sincronización completa: {e}")
    
    return qr_generados, errores

# Decorador para requerir login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debe iniciar sesión para acceder a esta página', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorador para requerir permisos de administrador
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debe iniciar sesión para acceder a esta página', 'error')
            return redirect(url_for('login'))
        if session.get('user_profile') != 'administrador':
            flash('No tienes permisos para acceder a esta página. Solo los administradores pueden gestionar el sistema.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Configuración para producción
app.config['DEBUG'] = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'

def crear_datos_maestros():
    """Crea datos maestros por defecto si no existen"""
    # Tipos de equipos
    tipos_equipos = [
        'Grúa Torre', 'Grúa Móvil', 'Montacargas', 'Excavadora', 
        'Bulldozer', 'Cargador', 'Grúa Portuaria', 'Grúa Overhead'
    ]
    
    for descripcion in tipos_equipos:
        if not TipoEquipo.query.filter_by(descripcion=descripcion).first():
            tipo = TipoEquipo(descripcion=descripcion)
            db.session.add(tipo)
    
    # Marcas
    marcas = [
        'Caterpillar', 'Liebherr', 'Terex', 'Manitowoc', 'Grove',
        'Kato', 'Kobelco', 'Hitachi', 'Komatsu', 'Volvo'
    ]
    
    for descripcion in marcas:
        if not Marca.query.filter_by(DescripcionMarca=descripcion).first():
            marca = Marca(DescripcionMarca=descripcion)
            db.session.add(marca)
    
    # Estados de equipos
    estados_equipos = [
        'Operativo', 'Mantenimiento', 'Averiado', 'Fuera de Servicio',
        'En Reparación', 'Stand By', 'Reservado'
    ]
    
    for descripcion in estados_equipos:
        if not EstadoEquipo.query.filter_by(Descripcion=descripcion).first():
            estado = EstadoEquipo(Descripcion=descripcion)
            db.session.add(estado)
    
    # Cargos
    cargos = [
        'Operador de Grúa', 'Supervisor de Operaciones', 'Mecánico', 
        'Jefe de Mantenimiento', 'Coordinador de Proyectos', 'Técnico de Seguridad',
        'Administrador de Flota', 'Capataz', 'Auxiliar de Operaciones'
    ]
    
    for descripcion in cargos:
        if not Cargo.query.filter_by(descripcionCargo=descripcion).first():
            cargo = Cargo(descripcionCargo=descripcion)
            db.session.add(cargo)
    
    db.session.commit()
    print("Datos maestros creados exitosamente")

# Crear tablas si no existen
with app.app_context():
    db.create_all()
    
    # Usuario administrador ya creado por script de inicialización
    # No crear automáticamente para evitar conflictos
    
    # Datos maestros ya creados por script de inicialización
    # No crear automáticamente para evitar conflictos
    # crear_datos_maestros()

@app.route('/')
def index():
    """Página principal - redirige al login si no está autenticado"""
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember = request.form.get('remember')
        next_page = request.form.get('next')  # Obtener la página de destino
        
        # Buscar usuario por documento o email
        user = User.query.filter(
            (User.documento == username) | (User.email == username)
        ).first()
        
        if user and user.check_password(password) and user.estado == 'activo':
            session['user'] = user.documento
            session['user_id'] = user.id
            session['user_name'] = user.nombre
            session['user_profile'] = user.perfil_usuario
            
            # Actualizar último login
            user.update_last_login()
            
            if remember:
                session.permanent = True
            flash(f'¡Bienvenido {user.nombre}! Has iniciado sesión correctamente.', 'success')
            
            # Redirigir a la página de destino si existe, sino al dashboard
            if next_page:
                # Si la URL es completa, extraer solo la ruta
                if next_page.startswith('http'):
                    from urllib.parse import urlparse
                    parsed_url = urlparse(next_page)
                    next_page = parsed_url.path
                
                return redirect(next_page)
            else:
                return redirect(url_for('dashboard'))
        else:
            flash('Usuario, contraseña incorrectos o cuenta inactiva. Inténtalo de nuevo.', 'error')
    
    # Obtener la página de destino desde los parámetros GET
    next_page = request.args.get('next')
    return render_template('login.html', next_page=next_page)

@app.route('/dashboard')
def dashboard():
    """Panel de control principal"""
    if 'user' not in session:
        flash('Debes iniciar sesión para acceder al panel.', 'error')
        return redirect(url_for('login'))
    
    # Obtener estadísticas de usuarios
    total_users = User.query.count()
    active_users = User.query.filter_by(estado='activo').count()
    admin_users = User.query.filter_by(perfil_usuario='administrador').count()
    employee_users = User.query.filter_by(perfil_usuario='empleado').count()
    
    # Obtener estadísticas de equipos
    total_equipos = Equipo.query.count()
    equipos_activos = Equipo.query.filter_by(Estado='activo').count()
    equipos_operando = 0
    equipos_disponibles = 0
    
    # Calcular equipos operando y disponibles
    equipos = Equipo.query.filter_by(Estado='activo').all()
    print(f"DEBUG DASHBOARD: Total equipos activos: {len(equipos)}")
    
    for equipo in equipos:
        esta_operando = equipo.esta_operando()
        print(f"DEBUG DASHBOARD: Equipo {equipo.Placa} - Operando: {esta_operando}")
        if esta_operando:
            equipos_operando += 1
        else:
            equipos_disponibles += 1
    
    
    print(f"DEBUG DASHBOARD: Equipos operando: {equipos_operando}, Disponibles: {equipos_disponibles}")
    
    # Obtener estado de empleados (solo empleados, no administradores)
    empleados_activos = User.query.filter_by(estado='activo', perfil_usuario='empleado').all()
    empleados_estado = []
    
    for empleado in empleados_activos:
        # Verificar si el empleado tiene una entrada sin salida
        # Buscar la última entrada
        ultima_entrada = RegistroHoras.query.filter_by(
            IdEmpleado=empleado.id,
            TipoRegistro='entrada'
        ).order_by(RegistroHoras.FechaCreacion.desc()).first()
        
        entrada_pendiente = None
        if ultima_entrada:
            # Buscar si hay una salida posterior a esta entrada
            salida_correspondiente = RegistroHoras.query.filter(
                RegistroHoras.IdEmpleado == empleado.id,
                RegistroHoras.IdEquipo == ultima_entrada.IdEquipo,
                RegistroHoras.TipoRegistro == 'salida',
                RegistroHoras.FechaCreacion > ultima_entrada.FechaCreacion
            ).first()
            
            if not salida_correspondiente:
                entrada_pendiente = ultima_entrada
        
        if entrada_pendiente:
            # Empleado está trabajando
            equipo_trabajando = Equipo.query.get(entrada_pendiente.IdEquipo)
            empleados_estado.append({
                'empleado': empleado,
                'estado': 'trabajando',
                'equipo': equipo_trabajando,
                'entrada': entrada_pendiente
            })
        else:
            # Empleado está libre
            empleados_estado.append({
                'empleado': empleado,
                'estado': 'libre',
                'equipo': None,
                'entrada': None
            })
    
    print(f"DEBUG DASHBOARD: Total empleados: {len(empleados_activos)}")
    print(f"DEBUG DASHBOARD: Empleados trabajando: {len([e for e in empleados_estado if e['estado'] == 'trabajando'])}")
    print(f"DEBUG DASHBOARD: Empleados libres: {len([e for e in empleados_estado if e['estado'] == 'libre'])}")
    
    # Obtener estado detallado de equipos
    equipos_estado_detalle = []
    
    for equipo in equipos:
        # Siempre obtener todos los operadores actuales (empleados trabajando)
        operadores_actuales = equipo.obtener_operadores_actuales()
        
        # Determinar si está "trabajando" (solo si hay formularios de salida pendientes)
        esta_operando = equipo.esta_operando()
        
        if esta_operando:
            # Equipo está trabajando (hay formularios de salida pendientes)
            equipos_estado_detalle.append({
                'equipo': equipo,
                'estado': 'trabajando',
                'operadores': operadores_actuales,
                'cantidad_operadores': len(operadores_actuales)
            })
        else:
            # Equipo está quieto/disponible (no hay formularios de salida pendientes)
            # Pero aún puede mostrar empleados trabajando
            equipos_estado_detalle.append({
                'equipo': equipo,
                'estado': 'quieto',
                'operadores': operadores_actuales,
                'cantidad_operadores': len(operadores_actuales)
            })
    
    print(f"DEBUG DASHBOARD: Total equipos: {len(equipos)}")
    print(f"DEBUG DASHBOARD: Equipos trabajando: {len([e for e in equipos_estado_detalle if e['estado'] == 'trabajando'])}")
    print(f"DEBUG DASHBOARD: Equipos quietos: {len([e for e in equipos_estado_detalle if e['estado'] == 'quieto'])}")
    
    # Obtener registros recientes de entrada y salida
    registros_recientes = RegistroHoras.query.join(
        User, RegistroHoras.IdEmpleado == User.id
    ).join(
        Equipo, RegistroHoras.IdEquipo == Equipo.IdEquipo
    ).order_by(
        RegistroHoras.FechaCreacion.desc()
    ).limit(20).all()
    
    print(f"DEBUG DASHBOARD: Registros recientes: {len(registros_recientes)}")
    
    stats = {
        'total_users': total_users,
        'active_users': active_users,
        'admin_users': admin_users,
        'employee_users': employee_users,
        'total_equipos': total_equipos,
        'equipos_activos': equipos_activos,
        'equipos_operando': equipos_operando,
        'equipos_disponibles': equipos_disponibles,
        'empleados_estado': empleados_estado,
        'equipos_estado_detalle': equipos_estado_detalle,
        'registros_recientes': registros_recientes
    }
    
    # Debug logs
    print(f"DEBUG DASHBOARD: user_profile en sesión: {session.get('user_profile')}")
    print(f"DEBUG DASHBOARD: user_id en sesión: {session.get('user_id')}")
    print(f"DEBUG DASHBOARD: user en sesión: {session.get('user')}")
    
    return render_template('dashboard.html', 
                         username=session.get('user_name', session['user']),
                         user_profile=session.get('user_profile'),
                         stats=stats)

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.pop('user', None)
    session.pop('user_id', None)
    session.pop('user_name', None)
    session.pop('user_profile', None)
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('login'))

@app.route('/cambiar-contrasena', methods=['GET', 'POST'])
@login_required
def cambiar_contrasena():
    """Cambiar contraseña del usuario actual"""
    form = CambiarContrasenaForm()
    
    if form.validate_on_submit():
        user = User.query.get(session['user_id'])
        if user:
            # Verificar contraseña actual
            if user.check_password(form.contrasena_actual.data):
                # Actualizar la contraseña usando el método correcto
                user.set_password(form.nueva_contrasena.data)
                db.session.commit()
                
                flash('Contraseña cambiada exitosamente', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('La contraseña actual es incorrecta', 'error')
        else:
            flash('Error: Usuario no encontrado', 'error')
    
    return render_template('cambiar_contrasena.html', form=form)

# ===== RUTAS CRUD USUARIOS =====

def require_admin(f):
    """Decorador para requerir permisos de administrador"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debes iniciar sesión para acceder a esta página.', 'error')
            return redirect(url_for('login'))
        
        user = User.query.get(session.get('user_id'))
        if not user or user.perfil_usuario != 'administrador':
            flash('No tienes permisos para acceder a esta página. Solo los administradores pueden gestionar el sistema.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

@app.route('/usuarios')
@require_admin
def usuarios():
    """Lista de usuarios con filtros y búsqueda"""
    search_form = UserSearchForm()
    
    # Obtener parámetros de búsqueda
    search = request.args.get('search', '')
    perfil_filter = request.args.get('perfil_filter', '')
    estado_filter = request.args.get('estado_filter', '')
    page = request.args.get('page', 1, type=int)
    
    # Construir consulta
    query = User.query
    
    if search:
        query = query.filter(
            (User.nombre.contains(search)) |
            (User.documento.contains(search)) |
            (User.email.contains(search))
        )
    
    if perfil_filter:
        query = query.filter_by(perfil_usuario=perfil_filter)
    
    if estado_filter:
        query = query.filter_by(estado=estado_filter)
    
    # Paginación
    users = query.order_by(User.fecha_creacion.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    
    return render_template('usuarios/lista.html', 
                         users=users, 
                         search_form=search_form,
                         search=search,
                         perfil_filter=perfil_filter,
                         estado_filter=estado_filter)

@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_usuario():
    """Crear nuevo usuario"""
    form = UserForm()
    
    if form.validate_on_submit():
        try:
            user = User(
                tipo_documento=form.tipo_documento.data,
                documento=form.documento.data,
                nombre=form.nombre.data,
                email=form.email.data,
                celular=form.celular.data,
                contrasena=form.contrasena.data,
                perfil_usuario=form.perfil_usuario.data
            )
            
            # Aplicar estado si se especificó
            if form.estado.data != 'activo':
                user.estado = form.estado.data
                if form.estado.data == 'inactivo':
                    user.fecha_inactividad = get_colombia_datetime()
            
            db.session.add(user)
            db.session.commit()
            
            flash(f'Usuario {user.nombre} creado exitosamente.', 'success')
            return redirect(url_for('usuarios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el usuario: {str(e)}', 'error')
    
    return render_template('usuarios/formulario.html', 
                         form=form, 
                         titulo='Nuevo Usuario')

@app.route('/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_usuario(id):
    """Editar usuario existente"""
    user = User.query.get_or_404(id)
    form = UserEditForm(obj=user, id=id)
    
    if form.validate_on_submit():
        try:
            user.tipo_documento = form.tipo_documento.data
            user.documento = form.documento.data
            user.nombre = form.nombre.data
            user.email = form.email.data
            user.celular = form.celular.data
            user.perfil_usuario = form.perfil_usuario.data
            user.estado = form.estado.data
            
            # Actualizar contraseña si se proporcionó
            if form.contrasena.data:
                user.set_password(form.contrasena.data)
            
            # Manejar fecha de inactividad
            if form.estado.data == 'inactivo' and user.estado != 'inactivo':
                user.fecha_inactividad = get_colombia_datetime()
            elif form.estado.data == 'activo' and user.estado == 'inactivo':
                user.fecha_inactividad = None
            
            db.session.commit()
            
            flash(f'Usuario {user.nombre} actualizado exitosamente.', 'success')
            return redirect(url_for('usuarios'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el usuario: {str(e)}', 'error')
    
    return render_template('usuarios/formulario.html', 
                         form=form, 
                         titulo='Editar Usuario',
                         user=user)

@app.route('/usuarios/ver/<int:id>')
@require_admin
def ver_usuario(id):
    """Ver detalles de un usuario"""
    user = User.query.get_or_404(id)
    return render_template('usuarios/detalle.html', user=user)

@app.route('/usuarios/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_usuario(id):
    """Eliminar usuario (soft delete)"""
    user = User.query.get_or_404(id)
    
    try:
        # No permitir eliminar el usuario administrador principal
        if user.documento == 'admin':
            flash('No se puede eliminar el usuario administrador principal.', 'error')
            return redirect(url_for('usuarios'))
        
        user.deactivate_user()
        flash(f'Usuario {user.nombre} desactivado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al eliminar el usuario: {str(e)}', 'error')
    
    return redirect(url_for('usuarios'))

@app.route('/usuarios/activar/<int:id>', methods=['POST'])
@require_admin
def activar_usuario(id):
    """Activar usuario"""
    user = User.query.get_or_404(id)
    
    try:
        user.activate_user()
        flash(f'Usuario {user.nombre} activado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar el usuario: {str(e)}', 'error')
    
    return redirect(url_for('usuarios'))

@app.route('/api/usuarios')
@require_admin
def api_usuarios():
    """API para obtener usuarios en formato JSON"""
    users = User.query.all()
    return jsonify([user.to_dict() for user in users])

# ===== RUTAS CRUD EQUIPOS =====

@app.route('/equipos')
@require_admin
def equipos():
    """Lista de equipos con filtros y búsqueda"""
    search_form = EquipoSearchForm()
    
    # Obtener parámetros de búsqueda
    search = request.args.get('search', '')
    tipo_filter = request.args.get('tipo_filter', 0, type=int)
    marca_filter = request.args.get('marca_filter', 0, type=int)
    estado_filter = request.args.get('estado_filter', '')
    estado_equipo_filter = request.args.get('estado_equipo_filter', 0, type=int)
    page = request.args.get('page', 1, type=int)
    
    # Construir consulta
    query = Equipo.query
    
    if search:
        query = query.filter(
            (Equipo.Placa.contains(search)) |
            (Equipo.Modelo.contains(search)) |
            (Equipo.Referencia.contains(search))
        )
    
    if tipo_filter > 0:
        query = query.filter_by(IdTipoEquipo=tipo_filter)
    
    if marca_filter > 0:
        query = query.filter_by(IdMarca=marca_filter)
    
    if estado_filter:
        query = query.filter_by(Estado=estado_filter)
    
    if estado_equipo_filter > 0:
        query = query.filter_by(IdEstadoEquipo=estado_equipo_filter)
    
    # Paginación
    equipos = query.order_by(Equipo.FechaCreacion.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    
    return render_template('equipos/lista.html', 
                         equipos=equipos, 
                         search_form=search_form,
                         search=search,
                         tipo_filter=tipo_filter,
                         marca_filter=marca_filter,
                         estado_filter=estado_filter,
                         estado_equipo_filter=estado_equipo_filter)

@app.route('/equipos/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_equipo():
    """Crear nuevo equipo"""
    form = EquipoForm()
    
    if form.validate_on_submit():
        try:
            equipo = Equipo(
                IdTipoEquipo=form.IdTipoEquipo.data,
                Placa=form.Placa.data,
                Capacidad=form.Capacidad.data,
                IdMarca=form.IdMarca.data,
                IdEstadoEquipo=form.IdEstadoEquipo.data,
                UsuarioCreacion=session.get('user_id'),
                Referencia=form.Referencia.data,
                Color=form.Color.data,
                Modelo=form.Modelo.data,
                CentroCostos=form.CentroCostos.data
            )
            
            # Aplicar estado si se especificó
            if form.Estado.data != 'activo':
                equipo.Estado = form.Estado.data
                if form.Estado.data == 'inactivo':
                    equipo.FechaInactivacion = get_colombia_datetime()
                    equipo.UsuarioInactivacion = session.get('user_id')
            
            db.session.add(equipo)
            db.session.commit()
            
            # Generar código QR para el nuevo equipo
            try:
                qr_path = generar_qr_equipo(equipo.IdEquipo, equipo.Placa)
                if qr_path:
                    # Sincronizar con producción
                    filename = f"qr_equipo_{equipo.IdEquipo}_{equipo.Placa}.png"
                    sync_to_production(qr_path, filename)
                    flash(f'Equipo {equipo.Placa} creado exitosamente. Código QR generado y sincronizado.', 'success')
                else:
                    flash(f'Equipo {equipo.Placa} creado exitosamente. Error generando código QR.', 'warning')
            except Exception as qr_error:
                print(f"❌ Error generando QR para equipo {equipo.IdEquipo}: {str(qr_error)}")
                flash(f'Equipo {equipo.Placa} creado exitosamente. Error generando código QR: {str(qr_error)}', 'warning')
            return redirect(url_for('equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el equipo: {str(e)}', 'error')
    
    return render_template('equipos/formulario.html', 
                         form=form, 
                         titulo='Nuevo Equipo')

@app.route('/equipos/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_equipo(id):
    """Editar equipo existente"""
    equipo = Equipo.query.get_or_404(id)
    form = EquipoForm(obj=equipo, id=id)
    
    if form.validate_on_submit():
        try:
            equipo.IdTipoEquipo = form.IdTipoEquipo.data
            equipo.Placa = form.Placa.data
            equipo.Capacidad = form.Capacidad.data
            equipo.IdMarca = form.IdMarca.data
            equipo.Referencia = form.Referencia.data
            equipo.Color = form.Color.data
            equipo.Modelo = form.Modelo.data
            equipo.CentroCostos = form.CentroCostos.data
            equipo.IdEstadoEquipo = form.IdEstadoEquipo.data
            equipo.Estado = form.Estado.data
            
            # Manejar fecha de inactivación
            if form.Estado.data == 'inactivo' and equipo.Estado != 'inactivo':
                equipo.FechaInactivacion = get_colombia_datetime()
                equipo.UsuarioInactivacion = session.get('user_id')
            elif form.Estado.data == 'activo' and equipo.Estado == 'inactivo':
                equipo.FechaInactivacion = None
                equipo.UsuarioInactivacion = None
            
            db.session.commit()
            
            flash(f'Equipo {equipo.Placa} actualizado exitosamente.', 'success')
            return redirect(url_for('equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el equipo: {str(e)}', 'error')
    
    return render_template('equipos/formulario.html', 
                         form=form, 
                         titulo='Editar Equipo',
                         equipo=equipo)

@app.route('/equipos/ver/<int:id>')
@require_admin
def ver_equipo(id):
    """Ver detalles de un equipo"""
    equipo = Equipo.query.get_or_404(id)
    return render_template('equipos/detalle.html', equipo=equipo)

@app.route('/equipos/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_equipo(id):
    """Eliminar equipo (soft delete)"""
    equipo = Equipo.query.get_or_404(id)
    
    try:
        equipo.inactivar_equipo(session.get('user_id'))
        flash(f'Equipo {equipo.Placa} desactivado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al desactivar el equipo: {str(e)}', 'error')
    
    return redirect(url_for('equipos'))

@app.route('/equipos/activar/<int:id>', methods=['POST'])
@require_admin
def activar_equipo(id):
    """Activar equipo"""
    equipo = Equipo.query.get_or_404(id)
    
    try:
        equipo.activar_equipo()
        flash(f'Equipo {equipo.Placa} activado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar el equipo: {str(e)}', 'error')
    
    return redirect(url_for('equipos'))

@app.route('/api/equipos')
@require_admin
def api_equipos():
    """API para obtener equipos en formato JSON"""
    equipos = Equipo.query.all()
    return jsonify([equipo.to_dict() for equipo in equipos])

# ===== RUTAS AJAX PARA CREAR DATOS MAESTROS =====

@app.route('/api/crear-tipo-equipo', methods=['POST'])
@require_admin
def crear_tipo_equipo_ajax():
    """Crear nuevo tipo de equipo vía AJAX"""
    try:
        data = request.get_json()
        descripcion = data.get('descripcion', '').strip()
        
        if not descripcion:
            return jsonify({
                'success': False,
                'message': 'La descripción es obligatoria'
            }), 400
        
        # Verificar si ya existe
        if TipoEquipo.query.filter_by(descripcion=descripcion).first():
            return jsonify({
                'success': False,
                'message': 'Ya existe un tipo de equipo con esa descripción'
            }), 400
        
        # Crear nuevo tipo
        tipo = TipoEquipo(
            descripcion=descripcion,
            estado='activo'
        )
        
        db.session.add(tipo)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'id': tipo.IdTipoEquipo,
            'descripcion': tipo.descripcion,
            'message': 'Tipo de equipo creado exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al crear el tipo de equipo: {str(e)}'
        }), 500

@app.route('/api/crear-marca', methods=['POST'])
@require_admin
def crear_marca_ajax():
    """Crear nueva marca vía AJAX"""
    try:
        data = request.get_json()
        descripcion = data.get('descripcion', '').strip()
        
        if not descripcion:
            return jsonify({
                'success': False,
                'message': 'La descripción es obligatoria'
            }), 400
        
        # Verificar si ya existe
        if Marca.query.filter_by(DescripcionMarca=descripcion).first():
            return jsonify({
                'success': False,
                'message': 'Ya existe una marca con esa descripción'
            }), 400
        
        # Crear nueva marca
        marca = Marca(
            DescripcionMarca=descripcion,
            estado='activo'
        )
        
        db.session.add(marca)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'id': marca.IdMarca,
            'descripcion': marca.DescripcionMarca,
            'message': 'Marca creada exitosamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al crear la marca: {str(e)}'
        }), 500

# ===== RUTAS CRUD DATOS MAESTROS =====

# ===== DATOS MAESTROS =====

@app.route('/datos_maestros')
@require_admin
def datos_maestros():
    """Página principal de datos maestros"""
    return render_template('datos_maestros/index.html')

# ===== TIPOS DE EQUIPOS =====

@app.route('/tipos-equipos')
@require_admin
def tipos_equipos():
    """Lista de tipos de equipos"""
    tipos = TipoEquipo.query.order_by(TipoEquipo.descripcion).all()
    return render_template('datos_maestros/tipos_equipos.html', tipos=tipos)

@app.route('/tipos-equipos/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_tipo_equipo():
    """Crear nuevo tipo de equipo"""
    form = TipoEquipoForm()
    
    if form.validate_on_submit():
        try:
            tipo = TipoEquipo(
                descripcion=form.descripcion.data,
                estado=form.estado.data
            )
            db.session.add(tipo)
            db.session.commit()
            
            flash(f'Tipo de equipo "{tipo.descripcion}" creado exitosamente.', 'success')
            return redirect(url_for('tipos_equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el tipo de equipo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_tipo_equipo.html', 
                         form=form, 
                         titulo='Nuevo Tipo de Equipo')

@app.route('/tipos-equipos/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_tipo_equipo(id):
    """Editar tipo de equipo existente"""
    tipo = TipoEquipo.query.get_or_404(id)
    form = TipoEquipoForm(obj=tipo)
    
    if form.validate_on_submit():
        try:
            tipo.descripcion = form.descripcion.data
            tipo.estado = form.estado.data
            
            db.session.commit()
            
            flash(f'Tipo de equipo "{tipo.descripcion}" actualizado exitosamente.', 'success')
            return redirect(url_for('tipos_equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el tipo de equipo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_tipo_equipo.html', 
                         form=form, 
                         titulo='Editar Tipo de Equipo',
                         tipo=tipo)

@app.route('/tipos-equipos/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_tipo_equipo(id):
    """Eliminar tipo de equipo (soft delete)"""
    tipo = TipoEquipo.query.get_or_404(id)
    
    try:
        # Verificar si hay equipos usando este tipo
        equipos_count = Equipo.query.filter_by(IdTipoEquipo=id).count()
        if equipos_count > 0:
            flash(f'No se puede eliminar el tipo "{tipo.descripcion}" porque tiene {equipos_count} equipos asociados.', 'error')
            return redirect(url_for('tipos_equipos'))
        
        tipo.estado = 'inactivo'
        db.session.commit()
        flash(f'Tipo de equipo "{tipo.descripcion}" desactivado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al eliminar el tipo de equipo: {str(e)}', 'error')
    
    return redirect(url_for('tipos_equipos'))

@app.route('/tipos-equipos/activar/<int:id>', methods=['POST'])
@require_admin
def activar_tipo_equipo(id):
    """Activar tipo de equipo"""
    tipo = TipoEquipo.query.get_or_404(id)
    
    try:
        tipo.estado = 'activo'
        db.session.commit()
        flash(f'Tipo de equipo "{tipo.descripcion}" activado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar el tipo de equipo: {str(e)}', 'error')
    
    return redirect(url_for('tipos_equipos'))

# ===== MARCAS =====

@app.route('/marcas')
@require_admin
def marcas():
    """Lista de marcas"""
    marcas = Marca.query.order_by(Marca.DescripcionMarca).all()
    return render_template('datos_maestros/marcas.html', marcas=marcas)

@app.route('/marcas/nueva', methods=['GET', 'POST'])
@require_admin
def nueva_marca():
    """Crear nueva marca"""
    form = MarcaForm()
    
    if form.validate_on_submit():
        try:
            marca = Marca(
                DescripcionMarca=form.DescripcionMarca.data,
                estado=form.estado.data
            )
            db.session.add(marca)
            db.session.commit()
            
            flash(f'Marca "{marca.DescripcionMarca}" creada exitosamente.', 'success')
            return redirect(url_for('marcas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la marca: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_marca.html', 
                         form=form, 
                         titulo='Nueva Marca')

@app.route('/marcas/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_marca(id):
    """Editar marca existente"""
    marca = Marca.query.get_or_404(id)
    form = MarcaForm(obj=marca)
    
    if form.validate_on_submit():
        try:
            marca.DescripcionMarca = form.DescripcionMarca.data
            marca.estado = form.estado.data
            
            db.session.commit()
            
            flash(f'Marca "{marca.DescripcionMarca}" actualizada exitosamente.', 'success')
            return redirect(url_for('marcas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la marca: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_marca.html', 
                         form=form, 
                         titulo='Editar Marca',
                         marca=marca)

@app.route('/marcas/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_marca(id):
    """Eliminar marca (soft delete)"""
    marca = Marca.query.get_or_404(id)
    
    try:
        # Verificar si hay equipos usando esta marca
        equipos_count = Equipo.query.filter_by(IdMarca=id).count()
        if equipos_count > 0:
            flash(f'No se puede eliminar la marca "{marca.DescripcionMarca}" porque tiene {equipos_count} equipos asociados.', 'error')
            return redirect(url_for('marcas'))
        
        marca.estado = 'inactivo'
        db.session.commit()
        flash(f'Marca "{marca.DescripcionMarca}" desactivada exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al eliminar la marca: {str(e)}', 'error')
    
    return redirect(url_for('marcas'))

@app.route('/marcas/activar/<int:id>', methods=['POST'])
@require_admin
def activar_marca(id):
    """Activar marca"""
    marca = Marca.query.get_or_404(id)
    
    try:
        marca.estado = 'activo'
        db.session.commit()
        flash(f'Marca "{marca.DescripcionMarca}" activada exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar la marca: {str(e)}', 'error')
    
    return redirect(url_for('marcas'))

# ===== ESTADOS DE EQUIPOS =====

@app.route('/estados-equipos')
@require_admin
def estados_equipos():
    """Lista de estados de equipos"""
    estados = EstadoEquipo.query.order_by(EstadoEquipo.Descripcion).all()
    return render_template('datos_maestros/estados_equipos.html', estados=estados)

@app.route('/estados-equipos/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_estado_equipo():
    """Crear nuevo estado de equipo"""
    form = EstadoEquipoForm()
    
    if form.validate_on_submit():
        try:
            estado = EstadoEquipo(
                Descripcion=form.Descripcion.data,
                Estado=form.Estado.data
            )
            db.session.add(estado)
            db.session.commit()
            
            flash(f'Estado de equipo "{estado.Descripcion}" creado exitosamente.', 'success')
            return redirect(url_for('estados_equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el estado de equipo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_estado_equipo.html', 
                         form=form, 
                         titulo='Nuevo Estado de Equipo')

@app.route('/estados-equipos/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_estado_equipo(id):
    """Editar estado de equipo existente"""
    estado = EstadoEquipo.query.get_or_404(id)
    form = EstadoEquipoForm(obj=estado)
    
    if form.validate_on_submit():
        try:
            estado.Descripcion = form.Descripcion.data
            estado.Estado = form.Estado.data
            
            db.session.commit()
            
            flash(f'Estado de equipo "{estado.Descripcion}" actualizado exitosamente.', 'success')
            return redirect(url_for('estados_equipos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el estado de equipo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_estado_equipo.html', 
                         form=form, 
                         titulo='Editar Estado de Equipo',
                         estado=estado)

@app.route('/estados-equipos/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_estado_equipo(id):
    """Eliminar estado de equipo (soft delete)"""
    estado = EstadoEquipo.query.get_or_404(id)
    
    try:
        # Verificar si hay equipos usando este estado
        equipos_count = Equipo.query.filter_by(IdEstadoEquipo=id).count()
        if equipos_count > 0:
            flash(f'No se puede eliminar el estado "{estado.Descripcion}" porque tiene {equipos_count} equipos asociados.', 'error')
            return redirect(url_for('estados_equipos'))
        
        estado.Estado = 'inactivo'
        db.session.commit()
        flash(f'Estado de equipo "{estado.Descripcion}" desactivado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al eliminar el estado de equipo: {str(e)}', 'error')
    
    return redirect(url_for('estados_equipos'))

@app.route('/estados-equipos/activar/<int:id>', methods=['POST'])
@require_admin
def activar_estado_equipo(id):
    """Activar estado de equipo"""
    estado = EstadoEquipo.query.get_or_404(id)
    
    try:
        estado.Estado = 'activo'
        db.session.commit()
        flash(f'Estado de equipo "{estado.Descripcion}" activado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar el estado de equipo: {str(e)}', 'error')
    
    return redirect(url_for('estados_equipos'))

# ===== CARGOS =====

@app.route('/cargos')
@require_admin
def cargos():
    """Lista de cargos"""
    cargos = Cargo.query.order_by(Cargo.descripcionCargo).all()
    return render_template('datos_maestros/cargos.html', cargos=cargos)

@app.route('/cargos/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_cargo():
    """Crear nuevo cargo"""
    form = CargoForm()
    
    if form.validate_on_submit():
        try:
            cargo = Cargo(
                descripcionCargo=form.descripcionCargo.data,
                Estado=form.Estado.data
            )
            db.session.add(cargo)
            db.session.commit()
            
            flash(f'Cargo "{cargo.descripcionCargo}" creado exitosamente.', 'success')
            return redirect(url_for('cargos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el cargo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_cargo.html', 
                         form=form, 
                         titulo='Nuevo Cargo')

@app.route('/cargos/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_cargo(id):
    """Editar cargo existente"""
    cargo = Cargo.query.get_or_404(id)
    form = CargoForm(obj=cargo)
    
    if form.validate_on_submit():
        try:
            cargo.descripcionCargo = form.descripcionCargo.data
            cargo.Estado = form.Estado.data
            
            db.session.commit()
            
            flash(f'Cargo "{cargo.descripcionCargo}" actualizado exitosamente.', 'success')
            return redirect(url_for('cargos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el cargo: {str(e)}', 'error')
    
    return render_template('datos_maestros/formulario_cargo.html', 
                         form=form, 
                         titulo='Editar Cargo',
                         cargo=cargo)

@app.route('/cargos/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_cargo(id):
    """Eliminar cargo (soft delete)"""
    cargo = Cargo.query.get_or_404(id)
    
    try:
        # Verificar si hay usuarios usando este cargo (futuro)
        # usuarios_count = User.query.filter_by(IdCargo=id).count()
        # if usuarios_count > 0:
        #     flash(f'No se puede eliminar el cargo "{cargo.descripcionCargo}" porque tiene {usuarios_count} usuarios asociados.', 'error')
        #     return redirect(url_for('cargos'))
        
        cargo.Estado = 'inactivo'
        db.session.commit()
        flash(f'Cargo "{cargo.descripcionCargo}" desactivado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al eliminar el cargo: {str(e)}', 'error')
    
    return redirect(url_for('cargos'))

@app.route('/cargos/activar/<int:id>', methods=['POST'])
@require_admin
def activar_cargo(id):
    """Activar cargo"""
    cargo = Cargo.query.get_or_404(id)
    
    try:
        cargo.Estado = 'activo'
        db.session.commit()
        flash(f'Cargo "{cargo.descripcionCargo}" activado exitosamente.', 'success')
        
    except Exception as e:
        flash(f'Error al activar el cargo: {str(e)}', 'error')
    
    return redirect(url_for('cargos'))

# ===== QR EQUIPOS =====

@app.route('/qr-equipos')
@require_admin
def qr_equipos():
    """Sección de códigos QR para equipos"""
    equipos = Equipo.query.filter_by(Estado='activo').all()
    
    # Verificar si existen códigos QR para cada equipo
    equipos_con_qr = []
    for equipo in equipos:
        qr_filename = f"qr_equipo_{equipo.IdEquipo}_{equipo.Placa}.png"
        qr_path = os.path.join(app.static_folder, 'qr_codes', qr_filename)
        tiene_qr = os.path.exists(qr_path)
        
        equipos_con_qr.append({
            'equipo': equipo,
            'tiene_qr': tiene_qr,
            'qr_url': f"qr_codes/{qr_filename}" if tiene_qr else None,
            'url_registro': f"{os.environ.get('BASE_URL', request.url_root.rstrip('/'))}/equipo/{equipo.IdEquipo}"
        })
    
    return render_template('qr_equipos/index.html', equipos_con_qr=equipos_con_qr)

@app.route('/regenerar-qr')
@require_admin
def regenerar_qr():
    """Regenera todos los códigos QR"""
    try:
        qr_generados, errores = regenerar_todos_qr()
        if errores == 0:
            flash(f'Se generaron {qr_generados} códigos QR exitosamente.', 'success')
        else:
            flash(f'Se generaron {qr_generados} códigos QR. {errores} errores.', 'warning')
    except Exception as e:
        flash(f'Error regenerando códigos QR: {str(e)}', 'error')
    
    return redirect(url_for('qr_equipos'))

@app.route('/ver-qr-equipos')
@login_required
def ver_qr_equipos():
    """Vista de códigos QR para empleados (solo lectura)"""
    equipos = Equipo.query.filter_by(Estado='activo').all()
    
    # Verificar si existen códigos QR para cada equipo
    equipos_con_qr = []
    for equipo in equipos:
        qr_filename = f"qr_equipo_{equipo.IdEquipo}_{equipo.Placa}.png"
        qr_path = os.path.join(app.static_folder, 'qr_codes', qr_filename)
        tiene_qr = os.path.exists(qr_path)
        
        equipos_con_qr.append({
            'equipo': equipo,
            'tiene_qr': tiene_qr,
            'qr_url': f"qr_codes/{qr_filename}" if tiene_qr else None,
            'url_registro': f"{os.environ.get('BASE_URL', request.url_root.rstrip('/'))}/equipo/{equipo.IdEquipo}"
        })
    
    return render_template('qr_equipos/ver_empleados.html', equipos_con_qr=equipos_con_qr)

# ===== EQUIPOS PARA EMPLEADOS =====

@app.route('/equipos-disponibles')
@login_required
def equipos_disponibles():
    """Lista de equipos disponibles para empleados (solo lectura)"""
    # Obtener equipos activos
    equipos = Equipo.query.filter_by(Estado='activo').all()
    
    # Agregar información de estado operacional
    equipos_info = []
    for equipo in equipos:
        esta_operando = equipo.esta_operando()
        operador_actual = equipo.obtener_operador_actual() if esta_operando else None
        
        equipos_info.append({
            'equipo': equipo,
            'esta_operando': esta_operando,
            'operador_actual': operador_actual
        })
    
    return render_template('equipos/disponibles.html', equipos_info=equipos_info)

# ===== CLIENTES =====

@app.route('/clientes')
@require_admin
def clientes():
    """Lista de clientes"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Filtros de búsqueda
    busqueda = request.args.get('busqueda', '')
    estado = request.args.get('estado', '')
    
    query = Cliente.query
    
    if busqueda:
        query = query.filter(
            db.or_(
                Cliente.NombreCliente.contains(busqueda),
                Cliente.Nit.contains(busqueda)
            )
        )
    
    if estado:
        query = query.filter(Cliente.Estado == estado)
    
    clientes = query.order_by(Cliente.NombreCliente).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    search_form = ClienteSearchForm()
    search_form.busqueda.data = busqueda
    search_form.estado.data = estado
    
    return render_template('clientes/lista.html', 
                         clientes=clientes, 
                         search_form=search_form)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
@require_admin
def nuevo_cliente():
    """Crear nuevo cliente"""
    form = ClienteForm()
    
    if form.validate_on_submit():
        try:
            cliente = Cliente(
                NombreCliente=form.NombreCliente.data,
                Nit=form.Nit.data,
                UsuarioCrea=session['user_id'],
                Estado=form.Estado.data
            )
            db.session.add(cliente)
            db.session.commit()
            
            flash(f'Cliente "{cliente.NombreCliente}" creado exitosamente.', 'success')
            return redirect(url_for('clientes'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear el cliente: {str(e)}', 'error')
    
    return render_template('clientes/formulario.html', 
                         form=form, 
                         titulo='Nuevo Cliente')

@app.route('/clientes/editar/<int:id>', methods=['GET', 'POST'])
@require_admin
def editar_cliente(id):
    """Editar cliente existente"""
    cliente = Cliente.query.get_or_404(id)
    form = ClienteForm(obj=cliente)
    form.cliente_id = id  # Para validación de NIT único
    
    if form.validate_on_submit():
        try:
            cliente.NombreCliente = form.NombreCliente.data
            cliente.Nit = form.Nit.data
            cliente.Estado = form.Estado.data
            
            db.session.commit()
            
            flash(f'Cliente "{cliente.NombreCliente}" actualizado exitosamente.', 'success')
            return redirect(url_for('clientes'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar el cliente: {str(e)}', 'error')
    
    return render_template('clientes/formulario.html', 
                         form=form, 
                         titulo='Editar Cliente',
                         cliente=cliente)

@app.route('/clientes/ver/<int:id>')
@require_admin
def ver_cliente(id):
    """Ver detalles del cliente"""
    cliente = Cliente.query.get_or_404(id)
    return render_template('clientes/detalle.html', cliente=cliente)

@app.route('/clientes/eliminar/<int:id>', methods=['POST'])
@require_admin
def eliminar_cliente(id):
    """Eliminar cliente (soft delete)"""
    cliente = Cliente.query.get_or_404(id)
    
    try:
        if cliente.Estado == 'activo':
            cliente.inactivar(session['user_id'])
            flash(f'Cliente "{cliente.NombreCliente}" desactivado exitosamente.', 'success')
        else:
            cliente.activar()
            flash(f'Cliente "{cliente.NombreCliente}" activado exitosamente.', 'success')
        
        db.session.commit()
        
    except Exception as e:
        flash(f'Error al cambiar el estado del cliente: {str(e)}', 'error')
    
    return redirect(url_for('clientes'))

@app.route('/api/clientes')
@require_admin
def api_clientes():
    """API para obtener clientes (para select dinámicos)"""
    clientes = Cliente.query.filter_by(Estado='activo').order_by(Cliente.NombreCliente).all()
    return jsonify([cliente.to_dict() for cliente in clientes])

# ===== REGISTRO DE HORAS =====

@app.route('/equipo/<int:equipo_id>')
def panel_equipo(equipo_id):
    """Panel principal del equipo con opciones disponibles"""
    equipo = Equipo.query.get_or_404(equipo_id)
    
    # Verificar si el usuario está autenticado
    if 'user' not in session:
        # Redirigir al login con la página de destino
        return redirect(url_for('login', next=request.url))
    
    # Obtener el usuario logueado
    usuario = User.query.get(session['user_id'])
    
    # Obtener información adicional del equipo
    tipo_equipo = equipo.tipo_equipo.descripcion if equipo.tipo_equipo else 'N/A'
    marca = equipo.marca.DescripcionMarca if equipo.marca else 'N/A'
    estado_equipo = equipo.estado_equipo.Descripcion if equipo.estado_equipo else 'N/A'
    
    # Obtener estadísticas del equipo
    registros_hoy = RegistroHoras.query.filter(
        RegistroHoras.IdEquipo == equipo_id,
        db.func.date(RegistroHoras.FechaEmpleado) == db.func.current_date()
    ).count()
    
    registros_mes = RegistroHoras.query.filter(
        RegistroHoras.IdEquipo == equipo_id,
        db.extract('month', RegistroHoras.FechaEmpleado) == db.extract('month', db.func.current_date()),
        db.extract('year', RegistroHoras.FechaEmpleado) == db.extract('year', db.func.current_date())
    ).count()
    
    # Verificar si hay entrada pendiente (buscar registros sin salida)
    entrada_pendiente = None
    # Nota: El modelo RegistroHoras no tiene campo FechaSalida, 
    # por lo que no podemos verificar entradas pendientes de esta manera
    
    return render_template('equipos/panel.html', 
                         equipo=equipo,
                         usuario=usuario,
                         tipo_equipo=tipo_equipo,
                         marca=marca,
                         estado_equipo=estado_equipo,
                         registros_hoy=registros_hoy,
                         registros_mes=registros_mes,
                         entrada_pendiente=entrada_pendiente)

@app.route('/registro/<int:equipo_id>')
def registro_horas(equipo_id):
    """Página de registro de horas para una grúa específica"""
    equipo = Equipo.query.get_or_404(equipo_id)
    
    # Verificar si el usuario está autenticado
    if 'user' not in session:
        # Redirigir al login con la página de destino
        return redirect(url_for('login', next=request.url))
    
    # Obtener el usuario logueado
    empleado_id = session['user_id']
    empleado = User.query.get(empleado_id)
    
    # Verificar si el empleado ya tiene una entrada pendiente
    # Buscar la entrada más reciente para este equipo y empleado
    ultima_entrada = RegistroHoras.query.filter_by(
        IdEquipo=equipo_id,
        IdEmpleado=empleado_id,
        TipoRegistro='entrada'
    ).order_by(RegistroHoras.FechaCreacion.desc()).first()
    
    entrada_pendiente = None
    if ultima_entrada:
        # Verificar si esta entrada tiene una salida correspondiente
        salida_correspondiente = RegistroHoras.query.filter_by(
            IdEquipo=equipo_id,
            IdEmpleado=empleado_id,
            TipoRegistro='salida'
        ).filter(RegistroHoras.FechaCreacion > ultima_entrada.FechaCreacion).first()
        
        # Si no hay salida después de la última entrada, hay entrada pendiente
        if not salida_correspondiente:
            entrada_pendiente = ultima_entrada
    
    # Debug: verificar estado de entrada pendiente
    print(f"DEBUG: Equipo {equipo_id}, Empleado {empleado_id}")
    print(f"DEBUG: Última entrada: {ultima_entrada.IdRegistro if ultima_entrada else 'None'}")
    if ultima_entrada:
        salida_correspondiente = RegistroHoras.query.filter_by(
            IdEquipo=equipo_id,
            IdEmpleado=empleado_id,
            TipoRegistro='salida'
        ).filter(RegistroHoras.FechaCreacion > ultima_entrada.FechaCreacion).first()
        print(f"DEBUG: Salida correspondiente: {salida_correspondiente.IdRegistro if salida_correspondiente else 'None'}")
    print(f"DEBUG: Entrada pendiente encontrada: {entrada_pendiente is not None}")
    if entrada_pendiente:
        print(f"DEBUG: Entrada pendiente ID: {entrada_pendiente.IdRegistro}, Fecha: {entrada_pendiente.FechaEmpleado}")
    
    form = RegistroHorasForm()
    form.IdEquipo.data = str(equipo_id)
    form.IdEmpleado.data = str(empleado_id)
    form.TipoRegistro.data = 'salida' if entrada_pendiente else 'entrada'
    
    # Establecer fecha y hora actuales de Colombia como predeterminadas
    now_colombia = get_colombia_datetime()
    
    # Establecer fecha y hora actuales como predeterminadas
    form.FechaEmpleado.data = now_colombia.date()
    form.HoraEmpleado.data = now_colombia.time()
    
    # Forzar el renderizado del campo de tiempo con el valor correcto
    form.HoraEmpleado.render_kw = {
        'class': 'form-control',
        'type': 'time',
        'value': now_colombia.strftime('%H:%M')
    }
    
    # Si es salida, cargar datos de la entrada
    cargo_nombre = None
    cliente_nombre = None
    valores_entrada = None
    if entrada_pendiente:
        form.IdCargo.data = entrada_pendiente.IdCargo
        form.IdCargo.render_kw = {'readonly': True, 'style': 'background-color: #f8f9fa;'}
        # Pre-seleccionar el mismo cliente que se usó en la entrada
        form.IdCliente.data = entrada_pendiente.IdCliente
        form.IdCliente.render_kw = {'readonly': True, 'style': 'background-color: #f8f9fa;'}
        # El estado del equipo puede cambiar (se puede dañar), así que es editable
        form.IdEstadoEquipo.data = entrada_pendiente.IdEstadoEquipo
        # Obtener el nombre del cargo para mostrar
        cargo_nombre = entrada_pendiente.cargo.descripcionCargo if entrada_pendiente.cargo else 'N/A'
        # Obtener el nombre del cliente para mostrar
        cliente_nombre = entrada_pendiente.cliente.NombreCliente if entrada_pendiente.cliente else 'N/A'
        # Obtener valores de entrada para validación
        valores_entrada = {
            'kilometraje': entrada_pendiente.Kilometraje,
            'horometro': entrada_pendiente.Horometro,
            'fecha_entrada': entrada_pendiente.FechaEmpleado.strftime('%d/%m/%Y'),
            'hora_entrada': entrada_pendiente.HoraEmpleado.strftime('%H:%M')
        }
    
    return render_template('registro_horas/formulario.html', 
                         form=form, 
                         equipo=equipo, 
                         empleado=empleado,
                         entrada_pendiente=entrada_pendiente,
                         cargo_nombre=cargo_nombre,
                         cliente_nombre=cliente_nombre,
                         valores_entrada=valores_entrada)

@app.route('/registro/<int:equipo_id>/procesar', methods=['POST'])
@login_required
def procesar_registro_horas(equipo_id):
    """Procesar el registro de horas"""
    equipo = Equipo.query.get_or_404(equipo_id)
    empleado = User.query.get(session['user_id'])  # Obtener el empleado logueado
    form = RegistroHorasForm()
    
    # Debug: verificar datos del formulario
    print(f"DEBUG: Formulario válido: {form.validate_on_submit()}")
    print(f"DEBUG: Errores del formulario: {form.errors}")
    print(f"DEBUG: Datos del formulario: {request.form}")
    
    # Validaciones adicionales de fecha y hora
    if form.FechaEmpleado.data and form.HoraEmpleado.data:
        try:
            fecha_hora_ingresada = datetime.combine(form.FechaEmpleado.data, form.HoraEmpleado.data)
            fecha_actual = get_colombia_datetime()
            
            # Convertir fecha_hora_ingresada a timezone-aware para comparar
            colombia_tz = pytz.timezone('America/Bogota')
            fecha_hora_ingresada = colombia_tz.localize(fecha_hora_ingresada)
            
            # Validar que no sea una fecha futura
            if fecha_hora_ingresada > fecha_actual:
                flash('❌ Error: La fecha y hora no pueden ser futuras. Por favor, ingresa una fecha y hora válida.', 'error')
                # Mantener datos del formulario y mostrar errores
                return render_template('registro_horas/formulario.html', 
                                     equipo=equipo, 
                                     empleado=empleado, 
                                     form=form, 
                                     entrada_pendiente=None,
                                     cargo_nombre=None,
                                     cliente_nombre=None,
                                     valores_entrada=None)
        except Exception as e:
            print(f"ERROR en validación de fecha: {e}")
            flash('❌ Error en validación de fecha y hora. Intenta nuevamente.', 'error')
            return render_template('registro_horas/formulario.html', 
                                 equipo=equipo, 
                                 empleado=empleado, 
                                 form=form, 
                                 entrada_pendiente=None,
                                 cargo_nombre=None,
                                 cliente_nombre=None,
                                 valores_entrada=None)
        
        # Si es formulario de salida, validar contra la entrada
        if form.TipoRegistro.data == 'salida':
            # Buscar la entrada correspondiente
            entrada = RegistroHoras.query.filter(
                RegistroHoras.IdEmpleado == empleado.id,
                RegistroHoras.IdEquipo == equipo_id,
                RegistroHoras.TipoRegistro == 'entrada'
            ).order_by(RegistroHoras.FechaCreacion.desc()).first()
            
            if entrada:
                fecha_hora_entrada = datetime.combine(entrada.FechaEmpleado, entrada.HoraEmpleado)
                # Convertir fecha_hora_entrada a timezone-aware para comparar
                fecha_hora_entrada = colombia_tz.localize(fecha_hora_entrada)
                
                # Validar que la salida no sea anterior a la entrada
                if fecha_hora_ingresada < fecha_hora_entrada:
                    flash(f'❌ Error: La fecha y hora de salida no pueden ser anteriores a la entrada. Entrada registrada: {entrada.FechaEmpleado.strftime("%d/%m/%Y")} a las {entrada.HoraEmpleado.strftime("%H:%M")}', 'error')
                    # Mantener datos del formulario y mostrar errores
                    return render_template('registro_horas/formulario.html', 
                                         equipo=equipo, 
                                         empleado=empleado, 
                                         form=form, 
                                         entrada_pendiente=entrada,
                                         cargo_nombre=entrada.cargo.descripcionCargo if entrada.cargo else 'N/A',
                                         cliente_nombre=entrada.cliente.NombreCliente if entrada.cliente else 'N/A',
                                         valores_entrada={
                                             'kilometraje': entrada.Kilometraje,
                                             'horometro': entrada.Horometro,
                                             'fecha_entrada': entrada.FechaEmpleado.strftime('%d/%m/%Y'),
                                             'hora_entrada': entrada.HoraEmpleado.strftime('%H:%M')
                                         })
                
                # Validar que no exceda 24 horas de diferencia
                diferencia_horas = (fecha_hora_ingresada - fecha_hora_entrada).total_seconds() / 3600
                if diferencia_horas > 24:
                    flash(f'❌ Error: La diferencia entre entrada y salida no puede ser mayor a 24 horas. Diferencia actual: {diferencia_horas:.1f} horas. Por favor, verifica las fechas.', 'error')
                    # Mantener datos del formulario y mostrar errores
                    return render_template('registro_horas/formulario.html', 
                                         equipo=equipo, 
                                         empleado=empleado, 
                                         form=form, 
                                         entrada_pendiente=entrada,
                                         cargo_nombre=entrada.cargo.descripcionCargo if entrada.cargo else 'N/A',
                                         cliente_nombre=entrada.cliente.NombreCliente if entrada.cliente else 'N/A',
                                         valores_entrada={
                                             'kilometraje': entrada.Kilometraje,
                                             'horometro': entrada.Horometro,
                                             'fecha_entrada': entrada.FechaEmpleado.strftime('%d/%m/%Y'),
                                             'hora_entrada': entrada.HoraEmpleado.strftime('%H:%M')
                                         })
    
    if form.validate_on_submit():
        try:
            # Obtener ubicación del navegador
            ubicacion = request.form.get('ubicacion', '')
            latitud = request.form.get('latitud', type=float)
            longitud = request.form.get('longitud', type=float)
            
            # Procesar archivos subidos
            foto_kilometraje = None
            foto_horometro = None
            foto_grua = None
            
            if form.FotoKilometraje.data:
                foto_kilometraje = guardar_archivo(form.FotoKilometraje.data, 'kilometraje')
            
            if form.FotoHorometro.data:
                foto_horometro = guardar_archivo(form.FotoHorometro.data, 'horometro')
            
            if form.FotoGrua.data:
                foto_grua = guardar_archivo(form.FotoGrua.data, 'grua')
            
            # Validaciones específicas
            if form.TipoRegistro.data == 'salida':
                # Validar que los valores sean mayores o iguales a la entrada
                entrada = RegistroHoras.query.filter_by(
                    IdEquipo=equipo_id,
                    IdEmpleado=session['user_id'],  # Usar el ID del usuario logueado
                    TipoRegistro='entrada'
                ).first()
                
                if entrada and entrada.es_operador():
                    if form.Kilometraje.data and form.Kilometraje.data < entrada.Kilometraje:
                        flash(f'❌ Error: El kilometraje de salida ({form.Kilometraje.data}) debe ser mayor o igual al de entrada ({entrada.Kilometraje}). El vehículo no puede retroceder en kilometraje.', 'error')
                        # Mantener datos del formulario y mostrar errores
                        return render_template('registro_horas/formulario.html', 
                                             equipo=equipo, 
                                             empleado=empleado, 
                                             form=form, 
                                             entrada_pendiente=entrada,
                                             cargo_nombre=entrada.cargo.descripcionCargo if entrada.cargo else 'N/A',
                                             cliente_nombre=entrada.cliente.NombreCliente if entrada.cliente else 'N/A',
                                             valores_entrada={
                                                 'kilometraje': entrada.Kilometraje,
                                                 'horometro': entrada.Horometro,
                                                 'fecha_entrada': entrada.FechaEmpleado.strftime('%d/%m/%Y'),
                                                 'hora_entrada': entrada.HoraEmpleado.strftime('%H:%M')
                                             })
                    
                    if form.Horometro.data and form.Horometro.data < entrada.Horometro:
                        flash(f'❌ Error: El horómetro de salida ({form.Horometro.data}) debe ser mayor o igual al de entrada ({entrada.Horometro}). El horómetro no puede retroceder.', 'error')
                        # Mantener datos del formulario y mostrar errores
                        return render_template('registro_horas/formulario.html', 
                                             equipo=equipo, 
                                             empleado=empleado, 
                                             form=form, 
                                             entrada_pendiente=entrada,
                                             cargo_nombre=entrada.cargo.descripcionCargo if entrada.cargo else 'N/A',
                                             cliente_nombre=entrada.cliente.NombreCliente if entrada.cliente else 'N/A',
                                             valores_entrada={
                                                 'kilometraje': entrada.Kilometraje,
                                                 'horometro': entrada.Horometro,
                                                 'fecha_entrada': entrada.FechaEmpleado.strftime('%d/%m/%Y'),
                                                 'hora_entrada': entrada.HoraEmpleado.strftime('%H:%M')
                                             })
            
            # Crear registro
            registro = RegistroHoras(
                IdEquipo=equipo_id,
                IdEmpleado=session['user_id'],  # Usar el ID del usuario logueado
                IdCargo=form.IdCargo.data,
                IdCliente=form.IdCliente.data if form.IdCliente.data else None,
                IdEstadoEquipo=form.IdEstadoEquipo.data,
                FechaEmpleado=form.FechaEmpleado.data,
                HoraEmpleado=form.HoraEmpleado.data,
                Kilometraje=form.Kilometraje.data,
                Horometro=form.Horometro.data,
                FotoKilometraje=foto_kilometraje,
                FotoHorometro=foto_horometro,
                FotoGrua=foto_grua,
                Observacion=form.Observacion.data,
                Ubicacion=ubicacion,
                Latitud=latitud,
                Longitud=longitud,
                TipoRegistro=form.TipoRegistro.data
            )
            
            db.session.add(registro)
            db.session.commit()
            
            # Debug: verificar que se guardó
            print(f"DEBUG: Registro guardado - ID: {registro.IdRegistro}, Tipo: {form.TipoRegistro.data}, Equipo: {equipo_id}, Empleado: {session['user_id']}")
            
            tipo = "entrada" if form.TipoRegistro.data == 'entrada' else "salida"
            
            # Mensaje de éxito más detallado
            if form.TipoRegistro.data == 'entrada':
                flash(f'✅ Registro de ENTRADA guardado exitosamente para {equipo.Placa} - {empleado.nombre}', 'success')
            else:
                flash(f'✅ Registro de SALIDA guardado exitosamente para {equipo.Placa} - {empleado.nombre}. El equipo está ahora disponible.', 'success')
            
            # Redirigir a página de confirmación
            return redirect(url_for('confirmacion_registro', 
                                 equipo_id=equipo_id, 
                                 tipo=tipo, 
                                 empleado_nombre=empleado.nombre,
                                 equipo_placa=equipo.Placa))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar el registro: {str(e)}', 'error')
    else:
        # Si el formulario no es válido, mostrar errores
        for field, errors in form.errors.items():
            for error in errors:
                flash(f'Error en {field}: {error}', 'error')
    
    return redirect(url_for('registro_horas', equipo_id=equipo_id))

def guardar_archivo(archivo, prefijo):
    """Guardar archivo subido y retornar el nombre"""
    if archivo and archivo.filename:
        import uuid
        from werkzeug.utils import secure_filename
        
        # Generar nombre único
        extension = archivo.filename.rsplit('.', 1)[1].lower()
        nombre_archivo = f"{prefijo}_{uuid.uuid4().hex}.{extension}"
        
        # Crear directorio si no existe
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # Guardar archivo
        ruta_archivo = os.path.join(app.config['UPLOAD_FOLDER'], nombre_archivo)
        archivo.save(ruta_archivo)
        
        return nombre_archivo
    return None

@app.route('/confirmacion/<int:equipo_id>/<tipo>')
@login_required
def confirmacion_registro(equipo_id, tipo):
    """Página de confirmación después del registro"""
    equipo = Equipo.query.get_or_404(equipo_id)
    empleado_nombre = request.args.get('empleado_nombre', 'N/A')
    equipo_placa = request.args.get('equipo_placa', equipo.Placa)
    
    # Obtener el último registro para mostrar detalles
    ultimo_registro = RegistroHoras.query.filter_by(
        IdEquipo=equipo_id,
        IdEmpleado=session['user_id']
    ).order_by(RegistroHoras.FechaCreacion.desc()).first()
    
    return render_template('registro_horas/confirmacion.html',
                         equipo=equipo,
                         tipo=tipo,
                         empleado_nombre=empleado_nombre,
                         equipo_placa=equipo_placa,
                         ultimo_registro=ultimo_registro)

# ===== REPORTES =====

@app.route('/reportes')
@require_admin
def reportes():
    """Página principal de reportes"""
    return render_template('reportes/index.html')

@app.route('/revision-formularios')
@login_required
def revision_formularios():
    """Revisión de formularios con filtros"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', type=int)
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    tipo_registro = request.args.get('tipo_registro', '')
    equipo_id = request.args.get('equipo_id', type=int)
    
    # Construir consulta base con JOINs explícitos
    query = RegistroHoras.query.join(User, RegistroHoras.IdEmpleado == User.id).join(Equipo, RegistroHoras.IdEquipo == Equipo.IdEquipo).join(Cargo, RegistroHoras.IdCargo == Cargo.IdCargo)
    
    # Aplicar filtros
    if empleado_id:
        query = query.filter(RegistroHoras.IdEmpleado == empleado_id)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(RegistroHoras.FechaEmpleado >= fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(RegistroHoras.FechaEmpleado <= fecha_hasta_obj)
        except ValueError:
            pass
    
    if tipo_registro:
        query = query.filter(RegistroHoras.TipoRegistro == tipo_registro)
    
    if equipo_id:
        query = query.filter(RegistroHoras.IdEquipo == equipo_id)
    
    # Ordenar por fecha de creación (más recientes primero)
    query = query.order_by(RegistroHoras.FechaCreacion.desc())
    
    # Paginación
    page = request.args.get('page', 1, type=int)
    per_page = 20
    registros = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Obtener datos para filtros
    empleados = User.query.filter_by(estado='activo').order_by(User.nombre).all()
    equipos = Equipo.query.filter_by(Estado='activo').order_by(Equipo.Placa).all()
    
    return render_template('reportes/revision_formularios.html',
                         registros=registros,
                         empleados=empleados,
                         equipos=equipos,
                         empleado_id=empleado_id,
                         fecha_desde=fecha_desde,
                         fecha_hasta=fecha_hasta,
                         tipo_registro=tipo_registro,
                         equipo_id=equipo_id)

@app.route('/revision-formularios/<int:registro_id>')
@login_required
def detalle_formulario(registro_id):
    """Detalle de un formulario específico"""
    registro = RegistroHoras.query.get_or_404(registro_id)
    
    return render_template('reportes/detalle_formulario.html', registro=registro)

@app.route('/reportes/horas-empleado')
@require_admin
def reporte_horas_empleado():
    """Reporte de horas trabajadas por empleado con calendario"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', type=int)
    now_colombia = get_colombia_datetime()
    mes = request.args.get('mes', now_colombia.month, type=int)
    año = request.args.get('año', now_colombia.year, type=int)
    
    # Obtener lista de empleados activos
    empleados = User.query.filter_by(estado='activo', perfil_usuario='empleado').all()
    
    # Datos del calendario
    calendario_data = {}
    resumen = {
        'total_horas': 0,
        'dias_trabajados': 0,
        'dias_no_trabajados': 0,
        'empleado_nombre': 'Seleccione un empleado'
    }
    
    if empleado_id:
        empleado = User.query.get(empleado_id)
        if empleado:
            resumen['empleado_nombre'] = empleado.nombre
            
            # Obtener registros del empleado para el mes/año
            fecha_inicio = date(año, mes, 1)
            fecha_fin = date(año, mes, calendar.monthrange(año, mes)[1])
            
            registros = RegistroHoras.query.filter(
                RegistroHoras.IdEmpleado == empleado_id,
                RegistroHoras.FechaEmpleado >= fecha_inicio,
                RegistroHoras.FechaEmpleado <= fecha_fin
            ).order_by(RegistroHoras.FechaEmpleado).all()
            
            # Procesar registros para el calendario
            dias_trabajados = set()
            horas_por_dia = {}
            
            for registro in registros:
                dia = registro.FechaEmpleado.day
                dias_trabajados.add(dia)
                
                if registro.TipoRegistro == 'entrada':
                    # Calcular horas trabajadas si hay salida correspondiente
                    # Buscar salida del mismo día o del día siguiente
                    salida = RegistroHoras.query.filter(
                        RegistroHoras.IdEmpleado == empleado_id,
                        RegistroHoras.IdEquipo == registro.IdEquipo,
                        RegistroHoras.TipoRegistro == 'salida',
                        RegistroHoras.FechaEmpleado >= registro.FechaEmpleado,
                        RegistroHoras.FechaEmpleado <= registro.FechaEmpleado + timedelta(days=1)
                    ).order_by(RegistroHoras.FechaEmpleado, RegistroHoras.HoraEmpleado).first()
                    
                    if salida:
                        # Calcular diferencia de tiempo
                        entrada_datetime = datetime.combine(registro.FechaEmpleado, registro.HoraEmpleado)
                        salida_datetime = datetime.combine(salida.FechaEmpleado, salida.HoraEmpleado)
                        horas_trabajadas = (salida_datetime - entrada_datetime).total_seconds() / 3600
                        
                        if dia not in horas_por_dia:
                            horas_por_dia[dia] = 0
                        horas_por_dia[dia] += horas_trabajadas
            
            # Crear datos del calendario
            for dia in range(1, calendar.monthrange(año, mes)[1] + 1):
                calendario_data[dia] = {
                    'trabajo': dia in dias_trabajados,
                    'horas': round(horas_por_dia.get(dia, 0), 2)
                }
            
            # Calcular resumen
            resumen['total_horas'] = round(sum(horas_por_dia.values()), 2)
            resumen['dias_trabajados'] = len(dias_trabajados)
            resumen['dias_no_trabajados'] = calendar.monthrange(año, mes)[1] - len(dias_trabajados)
    
    # Generar calendario
    cal = calendar.monthcalendar(año, mes)
    nombre_mes = calendar.month_name[mes]
    
    return render_template('reportes/horas_empleado.html', 
                         empleados=empleados,
                         empleado_id=empleado_id,
                         mes=mes,
                         año=año,
                         calendario_data=calendario_data,
                         calendario=cal,
                         nombre_mes=nombre_mes,
                         resumen=resumen,
                         calendar=calendar)


@app.route('/reportes/horas-empleado/excel')
@require_admin
def exportar_horas_empleado_excel():
    """Exportar reporte de horas por empleado a Excel con detalles de cargos"""
    # Obtener parámetros de filtro
    empleado_id = request.args.get('empleado_id', type=int)
    now_colombia = get_colombia_datetime()
    mes = request.args.get('mes', now_colombia.month, type=int)
    año = request.args.get('año', now_colombia.year, type=int)
    
    if not empleado_id:
        flash('Debe seleccionar un empleado para exportar el reporte', 'error')
        return redirect(url_for('reporte_horas_empleado'))
    
    empleado = User.query.get(empleado_id)
    if not empleado:
        flash('Empleado no encontrado', 'error')
        return redirect(url_for('reporte_horas_empleado'))
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horas_{empleado.nombre}_{mes}_{año}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título principal
    ws.merge_cells('A1:H1')
    ws['A1'] = f"REPORTE DE HORAS TRABAJADAS - {empleado.nombre.upper()}"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws['A1'].alignment = center_alignment
    
    # Información del empleado
    ws['A3'] = "Empleado:"
    ws['B3'] = empleado.nombre
    ws['A4'] = "Documento:"
    ws['B4'] = empleado.documento
    ws['A5'] = "Período:"
    ws['B5'] = f"{calendar.month_name[mes]} {año}"
    
    # Encabezados de la tabla
    headers = [
        'Fecha', 'Equipo', 'Tipo Equipo', 'Cargo', 'Hora Entrada', 
        'Hora Salida', 'Horas Trabajadas', 'Observaciones'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Obtener registros del empleado para el mes/año
    fecha_inicio = date(año, mes, 1)
    fecha_fin = date(año, mes, calendar.monthrange(año, mes)[1])
    
    registros = RegistroHoras.query.filter(
        RegistroHoras.IdEmpleado == empleado_id,
        RegistroHoras.FechaEmpleado >= fecha_inicio,
        RegistroHoras.FechaEmpleado <= fecha_fin
    ).order_by(RegistroHoras.FechaEmpleado, RegistroHoras.HoraEmpleado).all()
    
    # Procesar registros
    row = 8
    total_horas = 0
    dias_trabajados = set()
    
    # Agrupar registros por fecha y equipo
    registros_por_fecha_equipo = {}
    for registro in registros:
        key = (registro.FechaEmpleado, registro.IdEquipo)
        if key not in registros_por_fecha_equipo:
            registros_por_fecha_equipo[key] = {'entrada': None, 'salida': None}
        
        if registro.TipoRegistro == 'entrada':
            registros_por_fecha_equipo[key]['entrada'] = registro
        else:
            registros_por_fecha_equipo[key]['salida'] = registro
    
    # Procesar cada grupo de registros
    for (fecha, equipo_id), registros_grupo in registros_por_fecha_equipo.items():
        entrada = registros_grupo['entrada']
        salida = registros_grupo['salida']
        
        if entrada:
            dias_trabajados.add(fecha.day)
            
            # Obtener información del equipo y cargo
            equipo = Equipo.query.get(equipo_id)
            cargo = Cargo.query.get(entrada.IdCargo) if entrada.IdCargo else None
            
            # Calcular horas trabajadas
            horas_trabajadas = 0
            hora_salida_str = "Sin salida"
            
            if salida:
                entrada_datetime = datetime.combine(entrada.FechaEmpleado, entrada.HoraEmpleado)
                salida_datetime = datetime.combine(salida.FechaEmpleado, salida.HoraEmpleado)
                horas_trabajadas = (salida_datetime - entrada_datetime).total_seconds() / 3600
                hora_salida_str = salida.HoraEmpleado.strftime('%H:%M')
                total_horas += horas_trabajadas
            
            # Escribir fila
            ws.cell(row=row, column=1, value=fecha.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=2, value=equipo.Placa if equipo else "N/A").border = border
            ws.cell(row=row, column=3, value=equipo.tipo_equipo.descripcion if equipo and equipo.tipo_equipo else "N/A").border = border
            ws.cell(row=row, column=4, value=cargo.descripcionCargo if cargo else "N/A").border = border
            ws.cell(row=row, column=5, value=entrada.HoraEmpleado.strftime('%H:%M')).border = border
            ws.cell(row=row, column=6, value=hora_salida_str).border = border
            ws.cell(row=row, column=7, value=round(horas_trabajadas, 2)).border = border
            ws.cell(row=row, column=8, value=entrada.Observacion or "").border = border
            
            # Aplicar alineación
            for col in range(1, 9):
                ws.cell(row=row, column=col).alignment = center_alignment
            
            row += 1
    
    # Resumen al final
    row += 2
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "RESUMEN"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="366092")
    ws[f'A{row}'].alignment = center_alignment
    
    row += 1
    ws[f'A{row}'] = "Total de Horas Trabajadas:"
    ws[f'B{row}'] = round(total_horas, 2)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Días Trabajados:"
    ws[f'B{row}'] = len(dias_trabajados)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Días No Trabajados:"
    ws[f'B{row}'] = calendar.monthrange(año, mes)[1] - len(dias_trabajados)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    # Ajustar ancho de columnas
    column_widths = [12, 20, 15, 15, 12, 12, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Crear respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo
    nombre_archivo = f"Reporte_Horas_{empleado.nombre.replace(' ', '_')}_{mes}_{año}.xlsx"
    
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo}"'}
    )

@app.route('/reportes/horas-equipo')
@require_admin
def reporte_horas_equipo():
    """Reporte de horas trabajadas por equipo con horómetro"""
    equipo_id = request.args.get('equipo_id', type=int)
    now_colombia = get_colombia_datetime()
    mes = request.args.get('mes', now_colombia.month, type=int)
    año = request.args.get('año', now_colombia.year, type=int)
    
    equipos = Equipo.query.filter_by(Estado='activo').all()
    
    calendario_data = {}
    resumen = {
        'total_horas': 0,
        'total_horometro': 0,
        'dias_trabajados': 0,
        'dias_no_trabajados': 0,
        'equipo_placa': 'Seleccione un equipo',
        'horometro_inicial': 0,
        'horometro_final': 0
    }
    
    if equipo_id:
        equipo = Equipo.query.get(equipo_id)
        if equipo:
            resumen['equipo_placa'] = f"{equipo.Placa} - {equipo.tipo_equipo.descripcion if equipo.tipo_equipo else 'N/A'}"
            
            fecha_inicio = date(año, mes, 1)
            fecha_fin = date(año, mes, calendar.monthrange(año, mes)[1])
            
            # Obtener registros del equipo en el mes
            registros = RegistroHoras.query.filter(
                RegistroHoras.IdEquipo == equipo_id,
                RegistroHoras.FechaEmpleado >= fecha_inicio,
                RegistroHoras.FechaEmpleado <= fecha_fin
            ).order_by(RegistroHoras.FechaEmpleado, RegistroHoras.HoraEmpleado).all()
            
            dias_trabajados = set()
            horas_por_dia = {}
            horometro_por_dia = {}
            
            # Obtener horómetro inicial del mes (último registro del mes anterior)
            ultimo_registro_anterior = RegistroHoras.query.filter(
                RegistroHoras.IdEquipo == equipo_id,
                RegistroHoras.FechaEmpleado < fecha_inicio,
                RegistroHoras.Horometro.isnot(None)
            ).order_by(RegistroHoras.FechaEmpleado.desc(), RegistroHoras.HoraEmpleado.desc()).first()
            
            horometro_inicial = ultimo_registro_anterior.Horometro if ultimo_registro_anterior else 0
            resumen['horometro_inicial'] = horometro_inicial
            
            # Procesar registros del mes
            for registro in registros:
                dia = registro.FechaEmpleado.day
                
                if registro.TipoRegistro == 'entrada':
                    # Buscar la salida correspondiente
                    salida = RegistroHoras.query.filter(
                        RegistroHoras.IdEquipo == equipo_id,
                        RegistroHoras.IdEmpleado == registro.IdEmpleado,
                        RegistroHoras.TipoRegistro == 'salida',
                        RegistroHoras.FechaEmpleado >= registro.FechaEmpleado,
                        RegistroHoras.FechaEmpleado <= registro.FechaEmpleado + timedelta(days=1)
                    ).order_by(RegistroHoras.FechaEmpleado, RegistroHoras.HoraEmpleado).first()
                    
                    if salida and registro.Horometro is not None and salida.Horometro is not None:
                        dias_trabajados.add(dia)
                        
                        # Calcular horas trabajadas
                        entrada_datetime = datetime.combine(registro.FechaEmpleado, registro.HoraEmpleado)
                        salida_datetime = datetime.combine(salida.FechaEmpleado, salida.HoraEmpleado)
                        horas_trabajadas = (salida_datetime - entrada_datetime).total_seconds() / 3600
                        
                        # Validar que las horas sean positivas y razonables (máximo 24 horas)
                        if horas_trabajadas > 0 and horas_trabajadas <= 24:
                            # Calcular incremento de horómetro
                            incremento_horometro = salida.Horometro - registro.Horometro
                            
                            # Validar que el incremento de horómetro sea positivo
                            if incremento_horometro >= 0:
                                if dia not in horas_por_dia:
                                    horas_por_dia[dia] = 0
                                    horometro_por_dia[dia] = 0
                                
                                horas_por_dia[dia] += horas_trabajadas
                                horometro_por_dia[dia] += incremento_horometro
            
            # Llenar calendario
            for dia in range(1, calendar.monthrange(año, mes)[1] + 1):
                calendario_data[dia] = {
                    'trabajo': dia in dias_trabajados,
                    'horas': round(horas_por_dia.get(dia, 0), 2),
                    'horometro': round(horometro_por_dia.get(dia, 0), 2)
                }
            
            # Calcular resumen
            resumen['total_horas'] = round(sum(horas_por_dia.values()), 2)
            resumen['total_horometro'] = round(sum(horometro_por_dia.values()), 2)
            resumen['dias_trabajados'] = len(dias_trabajados)
            resumen['dias_no_trabajados'] = calendar.monthrange(año, mes)[1] - len(dias_trabajados)
            
            # Obtener horómetro final del mes
            ultimo_registro_mes = RegistroHoras.query.filter(
                RegistroHoras.IdEquipo == equipo_id,
                RegistroHoras.FechaEmpleado <= fecha_fin,
                RegistroHoras.Horometro.isnot(None)
            ).order_by(RegistroHoras.FechaEmpleado.desc(), RegistroHoras.HoraEmpleado.desc()).first()
            
            resumen['horometro_final'] = ultimo_registro_mes.Horometro if ultimo_registro_mes else horometro_inicial
    
    cal = calendar.monthcalendar(año, mes)
    nombre_mes = calendar.month_name[mes]
    
    return render_template('reportes/horas_equipo.html', 
                         equipos=equipos,
                         equipo_id=equipo_id,
                         mes=mes,
                         año=año,
                         calendario_data=calendario_data,
                         calendario=cal,
                         nombre_mes=nombre_mes,
                         resumen=resumen,
                         calendar=calendar)


@app.route('/reportes/horas-equipo/excel')
@require_admin
def exportar_horas_equipo_excel():
    """Exportar reporte de horas por equipo a Excel con detalles de empleados y cargos"""
    # Obtener parámetros de filtro
    equipo_id = request.args.get('equipo_id', type=int)
    now_colombia = get_colombia_datetime()
    mes = request.args.get('mes', now_colombia.month, type=int)
    año = request.args.get('año', now_colombia.year, type=int)
    
    if not equipo_id:
        flash('Debe seleccionar un equipo para exportar el reporte', 'error')
        return redirect(url_for('reporte_horas_equipo'))
    
    equipo = Equipo.query.get(equipo_id)
    if not equipo:
        flash('Equipo no encontrado', 'error')
        return redirect(url_for('reporte_horas_equipo'))
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horas_{equipo.Placa}_{mes}_{año}"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Título principal
    ws.merge_cells('A1:J1')
    ws['A1'] = f"REPORTE DE HORAS POR EQUIPO - {equipo.Placa.upper()}"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws['A1'].alignment = center_alignment
    
    # Información del equipo
    ws['A3'] = "Equipo:"
    ws['B3'] = equipo.Placa
    ws['A4'] = "Tipo:"
    ws['B4'] = equipo.tipo_equipo.descripcion if equipo.tipo_equipo else 'N/A'
    ws['A5'] = "Capacidad:"
    ws['B5'] = f"{equipo.Capacidad} toneladas"
    ws['A6'] = "Período:"
    ws['B6'] = f"{calendar.month_name[mes]} {año}"
    
    # Encabezados de la tabla
    headers = [
        'Fecha', 'Empleado', 'Documento', 'Cargo', 'Hora Entrada', 
        'Hora Salida', 'Horas Trabajadas', 'Horómetro Entrada', 'Horómetro Salida', 'Observaciones'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment
    
    # Obtener registros del equipo para el mes/año
    fecha_inicio = date(año, mes, 1)
    fecha_fin = date(año, mes, calendar.monthrange(año, mes)[1])
    
    registros = RegistroHoras.query.filter(
        RegistroHoras.IdEquipo == equipo_id,
        RegistroHoras.FechaEmpleado >= fecha_inicio,
        RegistroHoras.FechaEmpleado <= fecha_fin
    ).order_by(RegistroHoras.FechaEmpleado, RegistroHoras.HoraEmpleado).all()
    
    # Procesar registros
    row = 9
    total_horas = 0
    total_horometro = 0
    dias_trabajados = set()
    
    # Agrupar registros por fecha y empleado
    registros_por_fecha_empleado = {}
    for registro in registros:
        key = (registro.FechaEmpleado, registro.IdEmpleado)
        if key not in registros_por_fecha_empleado:
            registros_por_fecha_empleado[key] = {'entrada': None, 'salida': None}
        
        if registro.TipoRegistro == 'entrada':
            registros_por_fecha_empleado[key]['entrada'] = registro
        else:
            registros_por_fecha_empleado[key]['salida'] = registro
    
    # Procesar cada grupo de registros
    for (fecha, empleado_id), registros_grupo in registros_por_fecha_empleado.items():
        entrada = registros_grupo['entrada']
        salida = registros_grupo['salida']
        
        if entrada:
            dias_trabajados.add(fecha.day)
            
            # Obtener información del empleado y cargo
            empleado = User.query.get(empleado_id)
            cargo = Cargo.query.get(entrada.IdCargo) if entrada.IdCargo else None
            
            # Calcular horas trabajadas y horómetro
            horas_trabajadas = 0
            hora_salida_str = "Sin salida"
            horometro_entrada = entrada.Horometro if entrada.Horometro else 0
            horometro_salida = 0
            incremento_horometro = 0
            
            if salida:
                entrada_datetime = datetime.combine(entrada.FechaEmpleado, entrada.HoraEmpleado)
                salida_datetime = datetime.combine(salida.FechaEmpleado, salida.HoraEmpleado)
                horas_trabajadas = (salida_datetime - entrada_datetime).total_seconds() / 3600
                hora_salida_str = salida.HoraEmpleado.strftime('%H:%M')
                
                if salida.Horometro:
                    horometro_salida = salida.Horometro
                    incremento_horometro = horometro_salida - horometro_entrada
                    total_horometro += incremento_horometro
                
                total_horas += horas_trabajadas
            
            # Escribir fila
            ws.cell(row=row, column=1, value=fecha.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=2, value=empleado.nombre if empleado else "N/A").border = border
            ws.cell(row=row, column=3, value=empleado.documento if empleado else "N/A").border = border
            ws.cell(row=row, column=4, value=cargo.descripcionCargo if cargo else "N/A").border = border
            ws.cell(row=row, column=5, value=entrada.HoraEmpleado.strftime('%H:%M')).border = border
            ws.cell(row=row, column=6, value=hora_salida_str).border = border
            ws.cell(row=row, column=7, value=round(horas_trabajadas, 2)).border = border
            ws.cell(row=row, column=8, value=round(horometro_entrada, 1)).border = border
            ws.cell(row=row, column=9, value=round(horometro_salida, 1) if horometro_salida > 0 else "N/A").border = border
            ws.cell(row=row, column=10, value=entrada.Observacion or "").border = border
            
            # Aplicar alineación
            for col in range(1, 11):
                ws.cell(row=row, column=col).alignment = center_alignment
            
            row += 1
    
    # Resumen al final
    row += 2
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = "RESUMEN"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="366092")
    ws[f'A{row}'].alignment = center_alignment
    
    row += 1
    ws[f'A{row}'] = "Total de Horas Trabajadas:"
    ws[f'B{row}'] = round(total_horas, 2)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total Horómetro:"
    ws[f'B{row}'] = round(total_horometro, 1)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Días Trabajados:"
    ws[f'B{row}'] = len(dias_trabajados)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Días Sin Trabajo:"
    ws[f'B{row}'] = calendar.monthrange(año, mes)[1] - len(dias_trabajados)
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    
    # Ajustar ancho de columnas
    column_widths = [12, 20, 15, 15, 12, 12, 15, 15, 15, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Crear respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nombre del archivo
    nombre_archivo = f"Reporte_Horas_Equipo_{equipo.Placa}_{mes}_{año}.xlsx"
    
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo}"'}
    )

@app.route('/reportes/registros-dinamicos')
@require_admin
def reporte_registros_dinamicos():
    """Reporte dinámico de registros de entrada y salida de empleados"""
    # Parámetros de filtro
    empleado_id = request.args.get('empleado_id', type=int)
    equipo_id = request.args.get('equipo_id', type=int)
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    tipo_registro = request.args.get('tipo_registro', '')
    
    # Obtener listas para filtros
    empleados = User.query.filter_by(estado='activo').all()
    equipos = Equipo.query.filter_by(Estado='activo').all()
    
    # Construir consulta base
    query = RegistroHoras.query.join(User, RegistroHoras.IdEmpleado == User.id).join(Equipo, RegistroHoras.IdEquipo == Equipo.IdEquipo)
    
    # Aplicar filtros
    if empleado_id:
        query = query.filter(RegistroHoras.IdEmpleado == empleado_id)
    if equipo_id:
        query = query.filter(RegistroHoras.IdEquipo == equipo_id)
    if tipo_registro:
        query = query.filter(RegistroHoras.TipoRegistro == tipo_registro)
    if fecha_inicio:
        query = query.filter(RegistroHoras.FechaEmpleado >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
    if fecha_fin:
        query = query.filter(RegistroHoras.FechaEmpleado <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
    
    # Ordenar por fecha y hora descendente (más recientes primero)
    registros = query.order_by(
        RegistroHoras.FechaEmpleado.desc(),
        RegistroHoras.HoraEmpleado.desc()
    ).limit(1000).all()  # Limitar a 1000 registros para rendimiento
    
    # Estadísticas rápidas
    total_registros = len(registros)
    entradas = len([r for r in registros if r.TipoRegistro == 'entrada'])
    salidas = len([r for r in registros if r.TipoRegistro == 'salida'])
    
    # Empleados únicos en los resultados
    empleados_unicos = list(set([r.empleado for r in registros]))
    equipos_unicos = list(set([r.equipo for r in registros]))
    
    return render_template('reportes/registros_dinamicos.html',
                         registros=registros,
                         empleados=empleados,
                         equipos=equipos,
                         empleados_unicos=empleados_unicos,
                         equipos_unicos=equipos_unicos,
                         empleado_id=empleado_id,
                         equipo_id=equipo_id,
                         fecha_inicio=fecha_inicio,
                         fecha_fin=fecha_fin,
                         tipo_registro=tipo_registro,
                         total_registros=total_registros,
                         entradas=entradas,
                         salidas=salidas)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=app.config['DEBUG'])
